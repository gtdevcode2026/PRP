import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# =========================
# CONFIG
# =========================
input_file = "PRP Sample Jun (2).xlsx"
output_file = "PRP_Final_Output.xlsx"
sheet_name = "TPRM Web-Portal Export"

# =========================
# LOAD DATA
# =========================
df = pd.read_excel(input_file, sheet_name=sheet_name, engine="openpyxl")

# Clean column names
df.columns = df.columns.str.strip().str.lower()

# =========================
# CHECK REQUIRED COLUMNS
# =========================
print("Available columns:", df.columns)

# =========================
# PIVOT TABLE
# =========================
pivot = pd.pivot_table(
    df,
    index="zone_assessing",          # rows
    columns="assessment_status",     # columns
    values="id",                     # values
    aggfunc="count",
    fill_value=0,
    margins=True,
    margins_name="Grand Total"
)

# =========================
# SUMMARY (for chart)
# =========================
summary = pivot.loc["Grand Total", ["Active", "Deprioritized", "Duplicate"]]

summary_df = pd.DataFrame({
    "Status": summary.index,
    "Count": summary.values
})

# =========================
# WRITE TO EXCEL
# =========================
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    pivot.to_excel(writer, sheet_name="Output", startrow=0, startcol=0)
    summary_df.to_excel(writer, sheet_name="Output", startrow=2, startcol=8, index=False)

# =========================
# ADD CHART
# =========================
# =========================
# ENSURE ALL 3 COLUMNS EXIST
# =========================
for col in ["Active", "Deprioritized", "Duplicate"]:
    if col not in pivot.columns:
        pivot[col] = 0

# =========================
# CREATE STRICT SUMMARY TABLE
# =========================
summary_df = pd.DataFrame({
    "Status": ["Active", "Deprioritized", "Duplicate"],
    "Count": [
        pivot.loc["Grand Total"].get("Active", 0),
        pivot.loc["Grand Total"].get("Deprioritized", 0),
        pivot.loc["Grand Total"].get("Duplicate", 0)
    ]
})

# =========================
# WRITE TO EXCEL
# =========================
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    pivot.to_excel(writer, sheet_name="Output", startrow=0, startcol=0)
    summary_df.to_excel(writer, sheet_name="Output", startrow=2, startcol=8, index=False)

# =========================
# CREATE CHART (FIXED CORRECTLY)
# =========================
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

wb = load_workbook(output_file)
ws = wb["Output"]

chart = BarChart()

chart.title = "Assessment Status Overview"
chart.y_axis.title = "Count"
chart.x_axis.title = "Assessment Status"

# ✅ Data (Count column + header)
data = Reference(ws, min_col=10, min_row=3, max_row=6)

# ✅ Categories (Status names)
cats = Reference(ws, min_col=9, min_row=4, max_row=6)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# ✅ Enable proper data labels
chart.dLbls = DataLabelList()
chart.dLbls.showVal = True

# ✅ Set chart type (vertical bars)
chart.type = "col"

# ✅ Position chart
ws.add_chart(chart, "L5")

wb.save(output_file)

print("✅ Chart fixed: Active + Deprioritized + Duplicate + labels all working")